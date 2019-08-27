"""
Python code to convert REL encoding to jpg/tif image.

SK
08/20/2019
"""

import os
import glob
import cv2
import numpy as np
from PIL import Image
from openpyxl import load_workbook

# This function takes a string in run-length encoded format, and returns a list of pixels
def RLencToMask(runs,image_shape):
    p1 = []  # Run-start pixel locations
    p2 = []  # Run-lengths

    # Separate run-lengths and pixel locations into seperate lists
    x = str(runs).split(' ')
    i = 0
    for m in x:
        if i % 2 == 0:
            p1.append(m)
        else:
            p2.append(m)
        i += 1

    # Get all absolute pixel values
    pixels = []
    for start, length in zip(p1, p2):
        i = 0
        length = int(length)
        pix = int(start)
        while i < length:
            pixels.append(pix)
            pix += 1
            i += 1

    mask_img = np.zeros(image_shape, dtype=np.uint8)

    mask_img = mask_img.flatten(order="F")

    mask_img[pixels] = 255

    mask_img = mask_img.reshape(image_shape, order='F')
    return mask_img


if __name__ == '__main__':

    rle_excel_file = r"L:\Research and Engineering\Data for Projects\DARPA Ultrasound\Data\heatmap_work\kaggle_utz\\train_masks.xlsx"
    work_book = load_workbook(rle_excel_file)
    sheet = work_book.active

    img_shape = (420,580)
    test_images_path = r"L:\Research and Engineering\Data for Projects\DARPA Ultrasound\Data\heatmap_work\kaggle_utz\test\test\\"

    """
    test_images = glob.glob(os.path.join(test_images_path,"*.tif"))

    for _image in test_images:
        img_data = cv2.imread(_image)

       print(img_data.shape)
    """
    cell_idx = 0

    while True:
        cell_idx+=1

        cell_data = sheet.cell(row=cell_idx, column=3)

        if cell_idx == 5637:
            break

        rle_run = cell_data.value

        if rle_run != None:
            image_mask = RLencToMask(rle_run,img_shape)




            cv2.imshow("dd",image_mask)
            cv2.waitKey()
            continue

            print(pixels)